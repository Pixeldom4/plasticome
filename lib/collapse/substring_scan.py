"""
Base: exact-substring scan of the joined dataset (as delivered).

What it does:
  1. Reads the 'joined' sheet (one row per unique aa_sequence).
  2. Normalizes each sequence (strip whitespace, uppercase) for matching.
  3. Finds every pair where the shorter sequence is a CONTIGUOUS EXACT
     substring of the longer one, and unions them into groups (union-find).
  4. In each group: shortest = 'core', longer members = 'superset',
     with delta_aa (length diff) and core_position (prefix/suffix/internal/chain).
  5. Writes a 'substring-scan' sheet in the same format as the original.

IMPORTANT LIMITATION: the match test is `exact_substring_match`, i.e. Python's
`core_key in superset_key` (exact contiguous containment). It therefore catches
only terminal additions/truncations and exact internal containment. It does NOT
catch point substitutions or internal indels. To catch those, run
`substring_scan_extended.py`, which reuses this pipeline with a near-match test.

This module is importable: `run(match=..., position_fn=...)` drives the whole
pipeline, so the extended scan can override just the match predicate.
Running it directly reproduces the original exact-substring behavior.
"""

import os
import re
from collections import defaultdict

import pandas as pd
# openpyxl is imported lazily inside write_sheet() so the pure pipeline
# functions (norm/build_groups/assemble_rows) can be reused without the Excel
# engine installed — e.g. by run_substring_scan_csv.py.

try:                       # YAML is nicer for a hand-edited config; JSON is a fallback.
    import yaml
except ImportError:
    yaml = None

WORKBOOK = 'plasticome-v1_join_retrieving-from-accession.xlsx'
CONFIG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config.yaml')

# Defaults used when config.yaml is absent or omits a key.
DEFAULT_CONFIG = {
    'workbook': WORKBOOK,
    'joined_sheet': 'joined',
    'exact': {'sheet_name': 'substring-scan'},
    'near': {'sheet_name': 'near-substring-scan', 'min_coverage': 0.92},
}


def load_config(path=CONFIG_PATH):
    """Load config.yaml (or a .json fallback) merged over DEFAULT_CONFIG.

    The two nested sections ('exact', 'near') are merged key-by-key so a config
    that overrides just `near.min_coverage` keeps the default sheet names.
    """
    import copy
    cfg = copy.deepcopy(DEFAULT_CONFIG)
    if path and os.path.exists(path):
        with open(path) as f:
            if yaml is not None:
                user = yaml.safe_load(f) or {}
            else:                       # pragma: no cover - only when PyYAML missing
                import json
                user = json.load(f)
        for k, v in user.items():
            if isinstance(v, dict) and isinstance(cfg.get(k), dict):
                cfg[k].update(v)
            else:
                cfg[k] = v
    return cfg


def norm(s):
    """Normalize a sequence for matching: drop all whitespace, uppercase."""
    return None if pd.isna(s) else re.sub(r'\s+', '', str(s)).upper()


def coalesce_join(*vals):
    """Merge several ';'-delimited fields into one de-duplicated ';' string."""
    seen, out = set(), []
    for v in vals:
        if pd.isna(v):
            continue
        for part in str(v).split(';'):
            part = part.strip()
            if part and part not in seen:
                seen.add(part); out.append(part)
    return ';'.join(out) if out else ''


# ---- 1. load unique-sequence records from the joined sheet ----
def load_records(workbook=WORKBOOK, sheet='joined'):
    """Read the joined sheet into a DataFrame of unique-sequence records."""
    j = pd.read_excel(workbook, sheet_name=sheet)
    recs = []
    for _, row in j.iterrows():
        key = norm(row['aa_sequence'])
        if not key:
            continue
        recs.append({
            'key': key,
            'seq': str(row['aa_sequence']).strip(),
            'len': len(key),
            'enzyme_name': coalesce_join(row.get('pv1_enzyme_name'), row.get('retr_enzyme_name')),
            'accessions': coalesce_join(row.get('pv1_pazy_accession'), row.get('pv1_uniprot_accession'),
                                        row.get('pv1_pdb_accession'), row.get('retr_accession')),
            'source': row['match_status'],
        })
    return pd.DataFrame(recs).drop_duplicates(subset='key').reset_index(drop=True)


# ---- 2. match predicate + position labelling for the EXACT scan ----
def exact_substring_match(short_key, long_key):
    """THE EXACT-SUBSTRING TEST: is the shorter key a contiguous substring?"""
    return short_key in long_key


def exact_position(core_key, s):
    """Where the core sits inside a superset, assuming exact containment."""
    if   s.startswith(core_key): return 'prefix'    # core at N-term, addition at C-term
    elif s.endswith(core_key):   return 'suffix'    # core at C-term, addition at N-term
    elif core_key in s:          return 'internal'
    else:                        return 'chain'      # contains a smaller member, not the core


# ---- 3. union-find over the (parameterized) substring relation ----
def build_groups(df, match=exact_substring_match):
    """Union sequences via a pairwise `match(short_key, long_key) -> bool`.

    Iterates shortest-first, so for every tested pair the first arg is the
    shorter (candidate sub-sequence) and the second the longer. Only pairs of
    differing length are tested. Returns components with >1 member, sorted by
    (size desc, core length asc).
    """
    n = len(df)
    keys = df['key'].tolist()
    lens = df['len'].tolist()

    parent = list(range(n))
    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]; x = parent[x]
        return x
    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb: parent[ra] = rb

    order = sorted(range(n), key=lambda i: lens[i])   # shortest first
    for a_pos in range(n):
        i = order[a_pos]
        ki = keys[i]
        for b_pos in range(a_pos + 1, n):
            jx = order[b_pos]
            if lens[jx] == lens[i]:
                continue                    # equal length & distinct -> not a substring
            if match(ki, keys[jx]):
                union(i, jx)

    comp = defaultdict(list)
    for i in range(n):
        comp[find(i)].append(i)
    groups = [m for m in comp.values() if len(m) > 1]
    core_of = lambda m: min(m, key=lambda i: lens[i])
    groups.sort(key=lambda m: (-len(m), lens[core_of(m)]))
    return groups


# ---- 4. assemble the output rows ----
def assemble_rows(df, groups, position_fn=exact_position):
    """Turn groups into the flat sheet: shortest member = core, rest = superset."""
    keys = df['key'].tolist()
    lens = df['len'].tolist()
    rows_out = []
    for gi, members in enumerate(groups, start=1):
        core_i = min(members, key=lambda i: lens[i])
        core_key, core_len = keys[core_i], lens[core_i]
        for i in sorted(members, key=lambda i: lens[i]):
            if i == core_i:
                role, position = 'core', ''
            else:
                role, position = 'superset', position_fn(core_key, keys[i])
            rows_out.append({
                'group': gi, 'role': role, 'enzyme_name': df['enzyme_name'][i],
                'accessions': df['accessions'][i], 'source': df['source'][i],
                'length_aa': lens[i], 'delta_aa': lens[i] - core_len,
                'core_position': position, 'aa_sequence': df['seq'][i],
                'change_or_comment': '', 'most_wildtype': '',
            })
    return pd.DataFrame(rows_out)


# ---- 5. write the scan sheet ----
def write_sheet(scan, workbook=WORKBOOK, sheet_name='substring-scan'):
    """Write `scan` to `sheet_name` in `workbook`, styled like the original."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb = openpyxl.load_workbook(workbook)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ARIAL = 'Arial'
    hdr = Font(name=ARIAL, bold=True, color='FFFFFF'); hdr_fill = PatternFill('solid', fgColor='2F5496')
    core_fill = PatternFill('solid', fgColor='E2EFDA'); input_fill = PatternFill('solid', fgColor='FFF2CC')
    cols = list(scan.columns)
    ws.append(cols)
    for c in ws[1]:
        c.font = hdr; c.fill = hdr_fill; c.alignment = Alignment(horizontal='center')
    for _, r in scan.iterrows():
        ws.append([r[c] if pd.notna(r[c]) else None for c in cols])
    role_idx = cols.index('role') + 1
    manual_idx = {cols.index('change_or_comment') + 1, cols.index('most_wildtype') + 1}
    for rr in range(2, ws.max_row + 1):
        is_core = ws.cell(row=rr, column=role_idx).value == 'core'
        for cc in range(1, ws.max_column + 1):
            cell = ws.cell(row=rr, column=cc); cell.font = Font(name=ARIAL)
            if cc in manual_idx:   cell.fill = input_fill
            elif is_core:          cell.fill = core_fill
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}1"
    widths = {'group':7,'role':10,'enzyme_name':22,'accessions':26,'source':18,'length_aa':10,
              'delta_aa':9,'core_position':13,'aa_sequence':60,'change_or_comment':30,'most_wildtype':13}
    for i, c in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 14)
    wb.save(workbook)


# ---- orchestrator ----
def run(match=exact_substring_match, position_fn=exact_position, *,
        workbook=None, sheet_name=None, config=None):
    """Load -> group -> assemble -> write, returning the scan DataFrame.

    Settings come from `config` (or config.yaml via load_config) unless
    overridden by the explicit `workbook` / `sheet_name` arguments.
    """
    cfg = config or load_config()
    workbook = workbook or cfg['workbook']
    sheet_name = sheet_name or cfg['exact']['sheet_name']
    df = load_records(workbook, cfg['joined_sheet'])
    groups = build_groups(df, match)
    scan = assemble_rows(df, groups, position_fn)
    write_sheet(scan, workbook, sheet_name)
    print(f'sheet={sheet_name} groups={scan.group.nunique()} involved={len(scan)} '
          f"cores={(scan.role=='core').sum()} supersets={(scan.role=='superset').sum()}")
    return scan


if __name__ == '__main__':
    run()
